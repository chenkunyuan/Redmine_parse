"""Export HYDRANT Field Issue Tracking issues to Excel.

Incremental by default: only appends new issues not already in the sheet.
Use --full to regenerate the entire sheet from scratch.
"""
import argparse
import json
import os
import re
import sys

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from redmine_parse.client import RedmineClient

sys.stdout.reconfigure(encoding="utf-8")

def _load_config():
    candidates = [
        os.path.join(os.path.dirname(__file__), "config.json"),
        "config.json",
    ]
    for path in candidates:
        if os.path.exists(path):
            try:
                with open(path, "r", encoding="utf-8") as f:
                    return json.load(f)
            except (json.JSONDecodeError, OSError):
                pass
    return {}

_config = _load_config()
REDMINE_URL = os.environ.get("REDMINE_URL") or _config.get("REDMINE_URL") or "https://redmine.sercomm.co.jp"
REDMINE_API_KEY = os.environ.get("REDMINE_API_KEY") or _config.get("REDMINE_API_KEY") or ""
PROJECT_ID = 141

OUTPUT_FILE = os.path.join(os.path.dirname(__file__), "HYDRANT_Field_Issues.xlsx")

# --- Styles ---
hfont = Font(name="Arial", bold=True, size=10, color="FFFFFF")
hfill = PatternFill("solid", fgColor="4472C4")
halign = Alignment(horizontal="center", vertical="center")
border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
bfont = Font(name="Arial", size=10)
balign = Alignment(vertical="top")
lfont = Font(name="Arial", size=10, color="0000FF", underline="single")

cat_fills = {
    "WIFI": PatternFill("solid", fgColor="E2EFDA"),
    "WAN/Connection": PatternFill("solid", fgColor="D9E2F3"),
    "Power/Reboot": PatternFill("solid", fgColor="FCE4D6"),
    "VoIP/TEL": PatternFill("solid", fgColor="E4DFEC"),
    "LED/Lamp": PatternFill("solid", fgColor="FFF2CC"),
    "LAN": PatternFill("solid", fgColor="DDEBF7"),
    "FW/Filter": PatternFill("solid", fgColor="F8CBAD"),
    "GUI/Setting": PatternFill("solid", fgColor="C6EFCE"),
    "HW Defect": PatternFill("solid", fgColor="FFC7CE"),
    "Other": PatternFill("solid", fgColor="D9D9D9"),
}

headers = ["Issue", "Subject", "Priority", "Status", "Category", "Reporter", "Report Time"]
col_widths = [26, 80, 10, 14, 18, 10, 14]


def categorize(issue):
    subject = (issue.get("subject") or "")
    sl = subject.lower()
    if any(kw in sl for kw in ["電源", "power", "reboot", "再起動", "shutdown", "起動", "power off", "restart"]):
        return "Power/Reboot"
    if any(kw in sl for kw in ["voip", "tel", "電話", "phone", "sip", "通話", "音声", "fax"]):
        return "VoIP/TEL"
    if any(kw in sl for kw in ["led", "ランプ", "点滅", "blinking", "lamp"]):
        return "LED/Lamp"
    if any(kw in sl for kw in ["wifi", "wi-fi", "wireless", "無線", "ssid", "wlan", "wps", "mesh", "メッシュ"]):
        return "WIFI"
    if any(kw in sl for kw in ["wan", "internet", "インターネット", "onu", "ont", "pppoe", "optical", "光", "接続", "connect", "network", "ネット", "link", "回線", "dhcp", "ppp", "ip "]):
        return "WAN/Connection"
    if any(kw in sl for kw in ["lan", "ethernet", "イーサネット", "port", "ポート", "cable", "ケーブル"]):
        return "LAN"
    if any(kw in sl for kw in ["firewall", "filter", "フィルタ", "遮断", "block", "acl", "nat", "dmz"]):
        return "FW/Filter"
    if any(kw in sl for kw in ["gui", "webui", "画面", "表示", "設定", "setting", "browser"]):
        return "GUI/Setting"
    if any(kw in sl for kw in ["故障", "不良", "defect", "malfunction", "damage", "破損", "焼損", "burn", "short", "ショート", "物理"]):
        return "HW Defect"
    dl = (issue.get("description") or "")[:300].lower()
    if any(kw in dl for kw in ["電源", "power shutdown", "reboot", "restart"]):
        return "Power/Reboot"
    if any(kw in dl for kw in ["電話", "phone", "voip"]):
        return "VoIP/TEL"
    if any(kw in dl for kw in ["wifi", "wi-fi", "ssid", "wireless"]):
        return "WIFI"
    if any(kw in dl for kw in ["ネット接続", "internet connection", "wan"]):
        return "WAN/Connection"
    return "Other"


def report_month(created_on):
    if created_on and len(created_on) >= 7:
        return created_on[:7]
    return ""


def get_existing_ids(filepath):
    """Extract issue IDs already in the spreadsheet."""
    if not os.path.exists(filepath):
        return set()
    wb = load_workbook(filepath, read_only=True)
    ws = wb.active
    ids = set()
    for row in ws.iter_rows(min_col=1, max_col=1, min_row=2, values_only=True):
        cell_val = row[0]
        if cell_val:
            m = re.search(r"#(\d+)", str(cell_val))
            if m:
                ids.add(int(m.group(1)))
    wb.close()
    return ids


def write_header(ws):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = hfont
        cell.fill = hfill
        cell.alignment = halign
        cell.border = border


def write_row(ws, row_num, issue):
    tracker_name = (issue.get("tracker") or {}).get("name", "")
    issue_id = issue.get("id")
    label = f"{tracker_name} #{issue_id}"
    url = f"https://redmine.sercomm.co.jp/issues/{issue_id}"
    cat = categorize(issue)

    values = [
        label,
        issue.get("subject"),
        (issue.get("priority") or {}).get("name"),
        (issue.get("status") or {}).get("name"),
        cat,
        "Field",
        report_month(issue.get("created_on")),
    ]
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row_num, column=col, value=val)
        cell.font = bfont
        cell.alignment = balign
        cell.border = border

    id_cell = ws.cell(row=row_num, column=1)
    id_cell.font = lfont
    id_cell.hyperlink = url

    cat_cell = ws.cell(row=row_num, column=5)
    cat_cell.fill = cat_fills.get(cat, PatternFill())


def apply_sheet_formatting(ws, total_rows):
    for col, w in enumerate(col_widths, 1):
        from openpyxl.utils import get_column_letter
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{chr(64 + len(headers))}{total_rows + 1}"


def build_full(client):
    issues = client.get_issues(PROJECT_ID)
    print(f"Fetched {len(issues)} issues (full rebuild)")

    wb = Workbook()
    ws = wb.active
    ws.title = "HYDRANT Field Issues"
    write_header(ws)

    for r, issue in enumerate(issues, 2):
        write_row(ws, r, issue)

    apply_sheet_formatting(ws, len(issues))
    wb.save(OUTPUT_FILE)
    print(f"Saved: {OUTPUT_FILE}  ({len(issues)} rows)")


def build_incremental(client):
    existing = get_existing_ids(OUTPUT_FILE)
    print(f"Existing IDs in sheet: {len(existing)}")

    issues = client.get_issues(PROJECT_ID)
    print(f"Fetched from Redmine: {len(issues)}")

    new_issues = [i for i in issues if i.get("id") not in existing]
    if not new_issues:
        print("No new issues to add.")
        return

    if os.path.exists(OUTPUT_FILE):
        wb = load_workbook(OUTPUT_FILE)
        ws = wb.active
        start_row = ws.max_row + 1
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "HYDRANT Field Issues"
        write_header(ws)
        start_row = 2

    for i, issue in enumerate(new_issues):
        write_row(ws, start_row + i, issue)

    total = (start_row - 1) + len(new_issues)
    apply_sheet_formatting(ws, total)
    wb.save(OUTPUT_FILE)
    print(f"Added {len(new_issues)} new issues. Total: {total} rows.")
    for i in new_issues[:5]:
        print(f"  + #{i['id']}: {i.get('subject','')[:80]}")
    if len(new_issues) > 5:
        print(f"  ... and {len(new_issues) - 5} more")


def main():
    parser = argparse.ArgumentParser(description="Export HYDRANT Field issues to Excel")
    parser.add_argument("--full", action="store_true", help="Full rebuild instead of incremental")
    args = parser.parse_args()

    client = RedmineClient(REDMINE_URL, REDMINE_API_KEY)

    if args.full:
        build_full(client)
    else:
        build_incremental(client)


if __name__ == "__main__":
    main()
